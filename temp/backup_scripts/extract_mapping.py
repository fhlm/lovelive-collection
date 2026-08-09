import openpyxl
import os
import sys
import json
from collections import defaultdict

excel_path = r'D:\Program\workspace\lovelive-collection\lovelive趴趴图鉴26.8.7.xlsx'
output_dir = r'D:\Program\workspace\lovelive-collection\temp'

# 创建输出目录
os.makedirs(output_dir, exist_ok=True)

print(f'Reading Excel file: {excel_path}')
print(f'File size: {os.path.getsize(excel_path)} bytes')

try:
    # 使用 read_only=False 以访问图片
    wb = openpyxl.load_workbook(excel_path, read_only=False)
    print(f'Sheet names: {wb.sheetnames}')
    
    all_mappings = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n=== Processing sheet: {sheet_name} ===')
        
        # 读取表头（角色名）
        headers = []
        for cell in ws[1]:
            if cell.value is not None:
                headers.append((cell.column, cell.value))
        print(f'Headers (column, name): {headers[:10]}...')
        
        # 读取尺寸列（第一列）和种类列（第二列）
        sizes = []
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
            size, kind = row[0], row[1]
            if size is not None:
                sizes.append((size, kind))
        print(f'Sizes: {sizes[:5]}...')
        
        # 收集图片信息
        images = []
        for img in ws._images:
            # 图片对象包含 anchor 属性，可能是单元格锚点
            if hasattr(img, 'anchor') and img.anchor:
                # 获取单元格位置
                if hasattr(img.anchor, '_from') and img.anchor._from:
                    from_ = img.anchor._from
                    row = from_.row + 1  # openpyxl 行从0开始
                    col = from_.col + 1  # 列从0开始
                    images.append((row, col, img))
        print(f'Found {len(images)} images in sheet')
        
        # 对于每个图片，尝试获取对应的尺寸和角色
        for row, col, img in images:
            # 查找对应的尺寸和角色
            size = None
            kind = None
            character = None
            
            # 查找尺寸（从第一列）
            for r, (s, k) in enumerate(sizes, start=2):
                if r == row:
                    size = s
                    kind = k
                    break
            
            # 查找角色名（从表头）
            for c, name in headers:
                if c == col:
                    character = name
                    break
            
            if size and character:
                # 构建映射
                mapping = {
                    'series': sheet_name,
                    'size': size,
                    'kind': kind,
                    'character_ja': character,
                    'row': row,
                    'col': col,
                    'has_image': True
                }
                all_mappings.append(mapping)
                print(f'  Row {row}, Col {col}: {character} - {size} ({kind})')
    
    wb.close()
    
    # 保存映射到 JSON 文件
    output_file = os.path.join(output_dir, 'excel_mapping.json')
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_mappings, f, ensure_ascii=False, indent=2)
    print(f'\nSaved {len(all_mappings)} mappings to {output_file}')
    
    # 统计每个系列的映射数
    series_counts = defaultdict(int)
    for m in all_mappings:
        series_counts[m['series']] += 1
    print('\nMapping counts by series:')
    for series, count in sorted(series_counts.items()):
        print(f'  {series}: {count}')
    
except Exception as e:
    print(f'Error: {e}')
    import traceback
    traceback.print_exc()
    sys.exit(1)