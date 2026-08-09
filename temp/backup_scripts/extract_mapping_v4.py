import openpyxl
import os
import sys
import json
from collections import defaultdict
import math

excel_path = r'D:\Program\workspace\lovelive-collection\lovelive趴趴图鉴26.8.7.xlsx'
output_dir = r'D:\Program\workspace\lovelive-collection\temp'
os.makedirs(output_dir, exist_ok=True)

# 单位转换：EMU 到像素（近似）
EMU_PER_PIXEL = 9525  # 1 像素 = 9525 EMU

def get_column_width_in_pixels(ws, col_letter):
    """获取列宽（像素）"""
    if col_letter in ws.column_dimensions:
        width_chars = ws.column_dimensions[col_letter].width
        if width_chars:
            # 近似转换：1 字符宽度 ≈ 7 像素
            return width_chars * 7
    # 默认列宽
    return 64  # 像素

def get_row_height_in_pixels(ws, row_num):
    """获取行高（像素）"""
    if row_num in ws.row_dimensions:
        height_chars = ws.row_dimensions[row_num].height
        if height_chars:
            return height_chars * 1.33  # 近似转换
    # 默认行高
    return 20  # 像素

def calculate_column_overlap(img_left, img_right, col_start, col_end):
    """计算图片在指定列范围内的重叠宽度"""
    overlap_left = max(img_left, col_start)
    overlap_right = min(img_right, col_end)
    if overlap_left < overlap_right:
        return overlap_right - overlap_left
    return 0

print(f'Reading Excel file: {excel_path}')
print(f'File size: {os.path.getsize(excel_path)} bytes')

try:
    wb = openpyxl.load_workbook(excel_path, read_only=False)
    print(f'Sheet names: {wb.sheetnames}')
    
    all_mappings = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n=== Processing sheet: {sheet_name} ===')
        
        # 获取表头（角色名）
        headers = {}  # col_num -> character_name
        for cell in ws[1]:
            if cell.value is not None:
                headers[cell.column] = cell.value
        print(f'Headers: {headers}')
        
        # 获取尺寸和种类
        sizes = {}  # row_num -> (size, kind)
        for row in ws.iter_rows(min_row=2, max_col=2, values_only=False):
            row_num = row[0].row
            size = row[0].value
            kind = row[1].value if len(row) > 1 else None
            if size is not None:
                sizes[row_num] = (size, kind)
        print(f'Sizes: {sizes}')
        
        # 获取列宽
        col_widths = {}
        for col_num in range(1, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col_num)
            col_widths[col_num] = get_column_width_in_pixels(ws, col_letter)
        
        # 处理图片
        images = ws._images
        print(f'Found {len(images)} images')
        
        # 按行分组图片
        images_by_row = defaultdict(list)
        for img in images:
            if hasattr(img, 'anchor') and img.anchor:
                anchor = img.anchor
                if hasattr(anchor, '_from') and anchor._from:
                    from_ = anchor._from
                    start_row = from_.row + 1  # openpyxl 行从0开始
                    images_by_row[start_row].append(img)
        
        # 对于每个尺寸行，处理图片
        for row_num, row_images in images_by_row.items():
            if row_num not in sizes:
                continue
            size, kind = sizes[row_num]
            print(f'\nProcessing row {row_num}: size={size}, kind={kind}, images={len(row_images)}')
            
            # 计算该行每个图片的列位置
            processed_images = []
            for img in row_images:
                if hasattr(img, 'anchor') and img.anchor:
                    anchor = img.anchor
                    if hasattr(anchor, '_from') and anchor._from:
                        from_ = anchor._from
                        start_col = from_.col + 1
                        row_off = from_.rowOff if hasattr(from_, 'rowOff') else 0
                        col_off = from_.colOff if hasattr(from_, 'colOff') else 0
                        
                        # 计算图片左上角位置（像素）
                        cell_left = 0
                        for c in range(1, start_col):
                            cell_left += col_widths.get(c, 64)
                        cell_top = 0
                        for r in range(1, row_num):
                            cell_top += get_row_height_in_pixels(ws, r)
                        
                        # 加上偏移量（转换为像素）
                        img_left = cell_left + (col_off / EMU_PER_PIXEL)
                        img_top = cell_top + (row_off / EMU_PER_PIXEL)
                        
                        # 计算图片右下角位置
                        img_width = img.width if hasattr(img, 'width') else 100
                        img_height = img.height if hasattr(img, 'height') else 100
                        img_right = img_left + img_width
                        img_bottom = img_top + img_height
                        
                        # 计算图片跨越的列范围
                        col_overlaps = {}
                        current_col_start = 0
                        for col_num in range(1, ws.max_column + 1):
                            col_width = col_widths.get(col_num, 64)
                            col_end = current_col_start + col_width
                            
                            # 计算重叠
                            overlap = calculate_column_overlap(img_left, img_right, current_col_start, col_end)
                            if overlap > 0:
                                col_overlaps[col_num] = overlap
                            
                            current_col_start = col_end
                        
                        # 找到占比最大的列
                        if col_overlaps:
                            total_overlap = sum(col_overlaps.values())
                            best_col = max(col_overlaps.items(), key=lambda x: x[1])[0]
                            best_ratio = col_overlaps[best_col] / total_overlap if total_overlap > 0 else 0
                            
                            # 获取对应的角色
                            character = headers.get(best_col, f'Unknown_{best_col}')
                            
                            processed_images.append({
                                'img': img,
                                'start_col': start_col,
                                'best_col': best_col,
                                'character': character,
                                'best_ratio': best_ratio,
                                'col_overlaps': col_overlaps
                            })
            
            # 按照列排序
            processed_images.sort(key=lambda x: x['best_col'])
            
            # 确保每个角色只出现一次（去重）
            used_characters = set()
            unique_images = []
            for img_info in processed_images:
                char = img_info['character']
                if char not in used_characters:
                    used_characters.add(char)
                    unique_images.append(img_info)
                else:
                    print(f'  Duplicate character {char} in row {row_num}, skipping')
            
            # 为每个唯一图片创建映射
            for img_info in unique_images:
                mapping = {
                    'series': sheet_name,
                    'character_ja': img_info['character'],
                    'size': size,
                    'kind': kind,
                    'start_row': row_num,
                    'start_col': img_info['start_col'],
                    'best_col': img_info['best_col'],
                    'best_ratio': round(img_info['best_ratio'], 3),
                    'col_overlaps': {k: round(v, 2) for k, v in img_info['col_overlaps'].items()},
                    'img_width': img_info['img'].width if hasattr(img_info['img'], 'width') else 100,
                    'img_height': img_info['img'].height if hasattr(img_info['img'], 'height') else 100
                }
                all_mappings.append(mapping)
    
    wb.close()
    
    # 保存映射
    output_file = os.path.join(output_dir, 'excel_mapping_v4.json')
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(all_mappings, f, ensure_ascii=False, indent=2)
    print(f'\nSaved {len(all_mappings)} mappings to {output_file}')
    
    # 统计每个系列的映射数
    series_counts = defaultdict(int)
    for m in all_mappings:
        series_counts[m['series']] += 1
    print('\nMapping counts by series:')
    for series, count in sorted(series_counts.items()):
        print(f'  {series}: {count}')
    
    # 检查 μ's 冬季校服系列 40cm(L) 的映射
    print('\nChecking μ\'s winter uniform 40cm(L):')
    mu_mappings = [m for m in all_mappings if m['series'] == "μ's" and m['size'] == '40cm(L)' and m['kind'] == '冬季校服系列（冬制服）']
    print(f'Found {len(mu_mappings)} mappings')
    for m in mu_mappings:
        print(f'  Character: {m["character_ja"]}, Col: {m["best_col"]}, Ratio: {m["best_ratio"]:.3f}')
    
except Exception as e:
    print(f'Error: {e}')
    import traceback
    traceback.print_exc()
    sys.exit(1)