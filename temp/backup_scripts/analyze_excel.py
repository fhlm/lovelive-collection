import openpyxl
import sys

excel_path = r'D:\Program\workspace\lovelive-collection\lovelive趴趴图鉴26.8.7.xlsx'

try:
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n=== Sheet: {sheet_name} ===')
        
        # 收集所有非空单元格
        non_empty_cells = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            for col_idx, value in enumerate(row, start=1):
                if value is not None:
                    non_empty_cells.append((row_idx, col_idx, value))
        
        print(f'Total non-empty cells: {len(non_empty_cells)}')
        
        # 打印前20个非空单元格
        print('First 20 non-empty cells:')
        for i, (r, c, v) in enumerate(non_empty_cells[:20]):
            print(f'  Row {r}, Col {c}: {repr(v)[:100]}')
        
        # 统计每行的非空单元格数
        row_counts = {}
        for r, c, v in non_empty_cells:
            row_counts[r] = row_counts.get(r, 0) + 1
        print(f'\nRow counts (row: count):')
        for r in sorted(row_counts.keys())[:10]:
            print(f'  Row {r}: {row_counts[r]}')
        
        # 统计每列的非空单元格数
        col_counts = {}
        for r, c, v in non_empty_cells:
            col_counts[c] = col_counts.get(c, 0) + 1
        print(f'\nColumn counts (col: count):')
        for c in sorted(col_counts.keys())[:15]:
            print(f'  Col {c}: {col_counts[c]}')
    
    wb.close()
    
except Exception as e:
    print(f'Error: {e}')
    sys.exit(1)