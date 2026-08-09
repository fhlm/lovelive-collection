import openpyxl
import sys
import os

excel_path = r'D:\Program\workspace\lovelive-collection\lovelive趴趴图鉴26.8.7.xlsx'
print(f'Reading Excel file: {excel_path}')
print(f'File size: {os.path.getsize(excel_path)} bytes')

try:
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    print(f'Sheet names: {wb.sheetnames}')
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n=== Sheet: {sheet_name} ===')
        
        # 打印前5行
        print('First 5 rows:')
        for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True)):
            print(f'Row {i+1}: {row}')
        
        # 打印表头（假设第一行是表头）
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        print(f'\nHeaders: {headers}')
        
        # 统计非空行数（通过迭代）
        non_empty_rows = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                non_empty_rows += 1
        print(f'Non-empty rows (excluding header): {non_empty_rows}')
    
    wb.close()
    
except Exception as e:
    print(f'Error reading Excel file: {e}')
    sys.exit(1)